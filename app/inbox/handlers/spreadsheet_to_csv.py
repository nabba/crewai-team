"""Spreadsheet → CSV inbox handler.

PROGRAM §46.7 (Q9.4). Opens .xlsx / .ods with the lightest possible
backend (openpyxl for xlsx; odfpy for ods if installed), reads the
first sheet, writes a CSV to ``workspace/notes/<stem>.csv``.

Conservative — first sheet only, max 50k rows. Beyond that the
operator should be working in a proper data tool, not the inbox.
"""
from __future__ import annotations

import csv
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

logger = logging.getLogger(__name__)


_MAX_ROWS = 50_000


def run(path: Path) -> str:
    ext = path.suffix.lower().lstrip(".")
    if ext == "xlsx":
        rows = _read_xlsx(path)
    elif ext == "ods":
        rows = _read_ods(path)
    else:
        raise RuntimeError(f"unsupported spreadsheet extension: .{ext}")

    if not rows:
        raise RuntimeError("spreadsheet produced no rows")

    dest = _notes_dir() / f"{path.stem}.csv"
    if dest.exists():
        stem = dest.stem
        i = 1
        while True:
            cand = _notes_dir() / f"{stem}.{i}.csv"
            if not cand.exists():
                dest = cand
                break
            i += 1
    with open(dest, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow(r)
    return f"sheet → {dest.name} ({len(rows)} rows)"


def _read_xlsx(path: Path) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            f"openpyxl not installed: {exc}"
        ) from exc
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"xlsx open failed: {exc}") from exc
    try:
        ws = wb.active
        rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= _MAX_ROWS:
                rows.append([f"… truncated at {_MAX_ROWS} rows"])
                break
        return rows
    finally:
        wb.close()


def _read_ods(path: Path) -> list[list]:
    try:
        from odf.opendocument import load as ods_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError as exc:
        raise RuntimeError(
            f"odfpy not installed: {exc}"
        ) from exc
    try:
        doc = ods_load(str(path))
    except Exception as exc:
        raise RuntimeError(f"ods open failed: {exc}") from exc
    tables = doc.spreadsheet.getElementsByType(Table)
    if not tables:
        return []
    rows: list[list] = []
    for tr in tables[0].getElementsByType(TableRow):
        cells = []
        for cell in tr.getElementsByType(TableCell):
            parts = []
            for p in cell.getElementsByType(P):
                parts.append(str(p))
            cells.append("".join(parts))
        rows.append(cells)
        if len(rows) >= _MAX_ROWS:
            rows.append([f"… truncated at {_MAX_ROWS} rows"])
            break
    return rows


def _notes_dir() -> Path:
    from app.paths import WORKSPACE_ROOT
    d = Path(os.getenv("INBOX_NOTES_DIR", str(WORKSPACE_ROOT / "notes")))
    d.mkdir(parents=True, exist_ok=True)
    return d
